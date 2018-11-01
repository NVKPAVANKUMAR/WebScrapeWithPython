import unittest
from openpyxl import load_workbook


class TestExcelParse(unittest.TestCase):
    def test_parse_excel(self):
        wb = load_workbook('dataEngine.xlsx')
        print(wb.sheetnames)
        ws = wb.active
        for row in ws.iter_rows():
            for cell in row:
                print(cell.value)
